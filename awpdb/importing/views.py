# PR2018-04-14
from django.contrib.auth.decorators import login_required # PR2018-04-01
from django.shortcuts import render, redirect #, get_object_or_404
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.generic import UpdateView, DeleteView, View, ListView, CreateView

from schools.models import Country, Department, Schooldefault, Country_log, Schooldefault_log
from subjects.models import Subjectdefault,  Level, Sector
from students.models import Birthcountry, Birthcity
from awpdb import functions as f

# PR2018-04-27
import logging
logger = logging.getLogger(__name__)

# PR2018-05-06
from django.utils.translation import activate, get_language_info, ugettext_lazy as _



# PR2018-04-27 import Excel file from "How to upload and process the Excel file in Django" http://thepythondjango.com/upload-process-excel-file-django/
import openpyxl
from tablib import Dataset, import_set


@method_decorator([login_required], name='dispatch')
class ImportSchooldefaultView(View):

    def get(self, request):
        return render(request, 'import_schoolcode.html')

    def post(self,request):

        # Note that request.FILES will only contain data if the request method was POST and
        # the <form> that posted the request has the attribute enctype="multipart/form-data".
        # Otherwise, request.FILES will be empty.

        # from https://docs.djangoproject.com/en/2.0/ref/files/uploads/#django.core.files.uploadedfile.UploadedFile.name
        uploadedfile = request.FILES["excel_file"]
        logger.debug('import_schoolcode_excel_view uploadedfile.name: ' + str(uploadedfile.name))

        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(uploadedfile)

        # PR2018-04-27 debug: DeprecationWarning: Call to deprecated function get_sheet_names (Use wb.sheetnames). Was:  ws_names = wb.get_sheet_names()
        ws_names = wb.sheetnames
        logger.debug('import_schoolcode_excel_view ws_names: ' + str(ws_names))

        worksheet = wb.worksheets[0]
        logger.debug('import_schoolcode_excel_view worksheet: ' + str(worksheet))

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            logger.debug('import_schoolcode_excel_view row_data: ' + str(row_data))

            #PR2018-04-28 debug: don't forget de brackets when creating an instance of the class
            schooldefault = Schooldefault()
            code = str(row_data[0])
            schooldefault.code = str(row_data[0])
            name = str(row_data[1])
            schooldefault.name = str(name)
            countrycode = code[:3]  # slice: [<start position zero-based>:<end position but not included>]

            logger.debug('import_schoolcode_excel_view countrycode: ' + str(countrycode))
            if countrycode == "SXM":
                country =  Country.objects.get(abbrev='SXM')
            else:
                country =  Country.objects.get(abbrev='CUR')
            schooldefault.country = country
            logger.debug('import_schoolcode_excel_view country: ' + str(country))

            if len(code) == 5:
                seq = code[3:]
                if seq == '00':
                    schooldefault.is_default = True

            schooldefault.modified_by = request.user
            schooldefault.modified_at = timezone.now()
            schooldefault.save(request=self.request)

            excel_data.append(row_data)

        return render(request, 'import_schoolcode.html', {"excel_data": excel_data})


@method_decorator([login_required], name='dispatch')
class ImportSubjectdefaultView(View):

    def get(self, request):
        # permission: user.is_authenticated AND user.is_role_system
        param = {'display_school': True, 'display_user': True, 'override_school': request.user.role_str}
        headerbar_param = f.get_headerbar_param(request, param)
        # render(request object, template name, [dictionary optional]) returns an HttpResponse of the template rendered with the given context.
        return render(request, 'import_subjectdefault.html', headerbar_param)

    def post(self,request):
        # Note that request.FILES will only contain data if the request method was POST and
        # the <form> that posted the request has the attribute enctype="multipart/form-data".
        # Otherwise, request.FILES will be empty.

        # from https://docs.djangoproject.com/en/2.0/ref/files/uploads/#django.core.files.uploadedfile.UploadedFile.name
        uploadedfile = request.FILES["excel_file"]
        logger.debug('ImportSubjectdefaultView uploadedfile.name: ' + str(uploadedfile.name))

        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(uploadedfile)

        # PR2018-04-27 debug: DeprecationWarning: Call to deprecated function get_sheet_names (Use wb.sheetnames). Was:  ws_names = wb.get_sheet_names()
        ws_names = wb.sheetnames
        logger.debug('ImportSubjectdefaultView ws_names: ' + str(ws_names))

        worksheet = wb.worksheets[0]
        logger.debug('ImportSubjectdefaultView worksheet: ' + str(worksheet))

        excel_data = list()
        logger.debug('ImportSubjectdefaultView excel_data: ' + str(excel_data))

        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()

            for cell in row:
                row_data.append(str(cell.value))

            logger.debug('row_data: ' + str(row_data))

            #PR2018-04-28 debug: don't forget de brackets when creating an instance of the class
            sd = Subjectdefault()

            sd.name = str(row_data[0])
            sd.abbrev = str(row_data[1])
            sd.sequence = int(row_data[2])

            # PR2018-08-06 country = request_user.country
            sd.country = request.user.country

            sd.modified_by = request.user
            sd.modified_at = timezone.now()
            sd.save(request=self.request)

            excel_data.append(row_data)

        return render(request, 'import_subjectdefault.html', {"excel_data": excel_data})


@method_decorator([login_required], name='dispatch')
class ImportDepartmentView(View):

    def get(self, request):
        # permission: user.is_authenticated AND user.is_role_system_perm_admin
        param = {'display_school': True, 'display_user': True, 'override_school': request.user.role_str}
        headerbar_param = f.get_headerbar_param(request, param)

        logger.debug('ImportDepartmentView headerbar_param: ' + str(headerbar_param))
        # render(request object, template name, [dictionary optional]) returns an HttpResponse of the template rendered with the given context.
        return render(request, 'import_department.html', headerbar_param)

    def post(self,request):
        uploadedfile = request.FILES["excel_file"]
        # logger.debug('ImportDepartmentView uploadedfile.name: ' + str(uploadedfile.name))

        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(uploadedfile)
        # PR2018-04-27 debug: DeprecationWarning: Call to deprecated function get_sheet_names (Use wb.sheetnames). Was:  ws_names = wb.get_sheet_names()
        ws_names = wb.sheetnames
        # logger.debug('ImportSubjectdefaultView ws_names: ' + str(ws_names))
        worksheet = wb.worksheets[0]
        # logger.debug('ImportSubjectdefaultView worksheet: ' + str(worksheet))
        excel_data = list()
        # logger.debug('ImportSubjectdefaultView excel_data: ' + str(excel_data))

        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            # logger.debug('row_data: ' + str(row_data))

            #PR2018-04-28 debug: don't forget de brackets when creating an instance of the class
            dep = Department()

            dep.name = str(row_data[0])
            dep.abbrev = str(row_data[1])

            dep.country = request.user.country
            dep.modified_by = request.user
            dep.modified_at = timezone.now()

            dep.save(request=self.request)

            #logger.debug('subjectdefault.id: ' + str(subjectdefault.id) +
            #             ' .name: ' + str(subjectdefault.name) + ' .abbrev: ' + str(subjectdefault.abbrev) + ' .sequence: ' + str(subjectdefault.sequence))

            excel_data.append(row_data)

        return render(request, 'import_department.html', {"excel_data": excel_data})


@method_decorator([login_required], name='dispatch')
class ImportLevelView(View):

    def get(self, request):
        # permission: user.is_authenticated AND user.is_role_system_perm_admin
        param = {'display_school': True, 'display_user': True, 'override_school': request.user.role_str}
        headerbar_param = f.get_headerbar_param(request, param)
        # render(request object, template name, [dictionary optional]) returns an HttpResponse of the template rendered with the given context.
        return render(request, 'import_level.html', headerbar_param)

    def post(self, request):
        uploadedfile = request.FILES["excel_file"]
        # logger.debug('ImportDepartmentView uploadedfile.name: ' + str(uploadedfile.name))

        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(uploadedfile)
        # PR2018-04-27 debug: DeprecationWarning: Call to deprecated function get_sheet_names (Use wb.sheetnames). Was:  ws_names = wb.get_sheet_names()
        ws_names = wb.sheetnames
        # logger.debug('ImportSubjectdefaultView ws_names: ' + str(ws_names))
        worksheet = wb.worksheets[0]
        # logger.debug('ImportSubjectdefaultView worksheet: ' + str(worksheet))
        excel_data = list()
        # logger.debug('ImportSubjectdefaultView excel_data: ' + str(excel_data))

        # iterating over the rows and
        # getting value from each cell in row
        self.sequence = 1
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                logger.debug('cell.value: ' + str(cell.value))
            logger.debug('row_data: ' + str(row_data))
            row_data.append(self.sequence)
            logger.debug('row_data: ' + str(row_data))
            self.sequence += 1

            # PR2018-04-28 debug: don't forget de brackets when creating an instance of the class
            level = Level()

            level.name = str(row_data[0])
            level.abbrev = str(row_data[1])
            level.sequence = str(row_data[2])

            level.country = request.user.country
            level.modified_by = request.user
            level.modified_at = timezone.now()

            level.save(request=self.request)

            excel_data.append(row_data)

        return render(request, 'import_level.html', {"excel_data": excel_data})



@method_decorator([login_required], name='dispatch')
class ImportSectorView(View):

    def get(self, request):
        # permission: user.is_authenticated AND user.is_role_system_perm_admin
        param = {'display_school': True}
        headerbar_param = f.get_headerbar_param(request, param)
        # render(request object, template name, [dictionary optional]) returns an HttpResponse of the template rendered with the given context.
        return render(request, 'import_sector.html', headerbar_param)

    def post(self, request):
        uploadedfile = request.FILES["excel_file"]
        # logger.debug('ImportDepartmentView uploadedfile.name: ' + str(uploadedfile.name))

        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(uploadedfile)
        # PR2018-04-27 debug: DeprecationWarning: Call to deprecated function get_sheet_names (Use wb.sheetnames). Was:  ws_names = wb.get_sheet_names()
        ws_names = wb.sheetnames
        # logger.debug('ImportSubjectdefaultView ws_names: ' + str(ws_names))
        worksheet = wb.worksheets[0]
        # logger.debug('ImportSubjectdefaultView worksheet: ' + str(worksheet))
        excel_data = list()
        # logger.debug('ImportSubjectdefaultView excel_data: ' + str(excel_data))

        # iterating over the rows and
        # getting value from each cell in row
        self.sequence = 1
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                logger.debug('cell.value: ' + str(cell.value))
            logger.debug('row_data: ' + str(row_data))
            row_data.append(self.sequence)
            logger.debug('row_data: ' + str(row_data))
            self.sequence += 1

            # PR2018-04-28 debug: don't forget de brackets when creating an instance of the class
            sector = Sector()

            sector.name = str(row_data[0])
            sector.abbrev = str(row_data[1])
            sector.sequence = str(row_data[2])

            sector.country = request.user.country
            sector.modified_by = request.user
            sector.modified_at = timezone.now()

            sector.save(request=self.request)

            excel_data.append(row_data)

        return render(request, 'import_sector.html', {"excel_data": excel_data})


@method_decorator([login_required], name='dispatch')
class ImportBirthcountryView(View):  # PR2018-08-31

    def get(self, request):
        # permission: user.is_authenticated AND user.is_role_system_perm_admin
        param = {'display_school': True, 'display_user': True, 'override_school': request.user.role_str}
        headerbar_param = f.get_headerbar_param(request, param)
        # render(request object, template name, [dictionary optional]) returns an HttpResponse of the template rendered with the given context.
        return render(request, 'import_birthcountry.html', headerbar_param)

    def post(self, request):
        uploadedfile = request.FILES["excel_file"]
        # logger.debug('ImportDepartmentView uploadedfile.name: ' + str(uploadedfile.name))

        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(uploadedfile)
        # PR2018-04-27 debug: DeprecationWarning: Call to deprecated function get_sheet_names (Use wb.sheetnames). Was:  ws_names = wb.get_sheet_names()
        ws_names = wb.sheetnames
        # logger.debug('ImportSubjectdefaultView ws_names: ' + str(ws_names))
        worksheet = wb.worksheets[0]
        # logger.debug('ImportSubjectdefaultView worksheet: ' + str(worksheet))
        excel_data = list()
        # logger.debug('ImportSubjectdefaultView excel_data: ' + str(excel_data))

        # iterating over the rows and
        # getting value from each cell in row
        self.sequence = 1
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                logger.debug('cell.value: ' + str(cell.value))
            logger.debug('row_data: ' + str(row_data))
            row_data.append(self.sequence)
            logger.debug('row_data: ' + str(row_data))
            self.sequence += 1

            #PR2018-04-28 debug: don't forget de brackets when creating an instance of the class
            birthcountry = Birthcountry()

            birthcountry.name = str(row_data[0])
            birthcountry.modified_by = request.user
            birthcountry.modified_at = timezone.now()

            birthcountry.save(request=self.request)

            excel_data.append(row_data)

        return render(request, 'import_birthcountry.html', {"excel_data": excel_data})



@method_decorator([login_required], name='dispatch')
class ImportBirthcityView(View):  # PR2018-09-01

    def get(self, request):
        # permission: user.is_authenticated AND user.is_role_system_perm_admin
        param = {'display_school': True, 'display_user': True, 'override_school': request.user.role_str}
        headerbar_param = f.get_headerbar_param(request, param)
        # render(request object, template name, [dictionary optional]) returns an HttpResponse of the template rendered with the given context.
        return render(request, 'import_birthcity.html', headerbar_param)

    def post(self, request):
        uploadedfile = request.FILES["excel_file"]
        # logger.debug('ImportDepartmentView uploadedfile.name: ' + str(uploadedfile.name))

        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(uploadedfile)
        # PR2018-04-27 debug: DeprecationWarning: Call to deprecated function get_sheet_names (Use wb.sheetnames). Was:  ws_names = wb.get_sheet_names()
        ws_names = wb.sheetnames
        # logger.debug('ImportSubjectdefaultView ws_names: ' + str(ws_names))
        worksheet = wb.worksheets[0]
        # logger.debug('ImportSubjectdefaultView worksheet: ' + str(worksheet))
        excel_data = list()
        # logger.debug('ImportSubjectdefaultView excel_data: ' + str(excel_data))

        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            # logger.debug('row_data: ' + str(row_data))

            #PR2018-04-28 debug: don't forget de brackets when creating an instance of the class
            birthcity = Birthcity()

            # get birthcountry
            countryname = str(row_data[0])
            if countryname is not None:
                logger.debug('countryname: ' + str(countryname) + ' Type: ' + str(type(countryname)))
                birthcountry = Birthcountry.objects.filter(name=countryname).first()
                logger.debug('birthcountry: ' + str(birthcountry) + ' Type: ' + str(type(birthcountry)))
                if birthcountry:

                    logger.debug('birthcountry: Type: ' + str(type(birthcountry)))

                    if birthcountry is not None:
                        logger.debug('birthcountry: ' + str(birthcountry) + ' Type: ' + str(type(birthcountry)))
                        birthcity.birthcountry = birthcountry
                        birthcity.name = str(row_data[1])
                        birthcity.modified_by = request.user
                        birthcity.modified_at = timezone.now()

                        birthcity.save(request=self.request)
                    else:
                        row_data[0] +=  " (not found)"
            else:
                row_data[0] = "no country"
            excel_data.append(row_data)

        return render(request, 'import_birthcity.html', {"excel_data": excel_data})
